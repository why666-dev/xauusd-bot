"""
excel_logger.py — Excel Trade Logger for XAUUSD Bot
Per-strategy sheets: All Trades | S1 Asian Breakout | S2 Goldmine | S3 Silver Bullet | Performance
"""

import os
import threading
import logging
from datetime import datetime
import pytz

log        = logging.getLogger(__name__)
ET         = pytz.timezone("America/New_York")
IST        = pytz.timezone("Asia/Kolkata")
EXCEL_PATH = "xauusd_trades.xlsx"
COMMISSION = 0.005  # approximate per oz
_lock      = threading.Lock()

C = {
    "card":    "060B14", "alt":     "0C1420", "border":  "1E3048",
    "gold":    "FFB800", "green":   "00E676", "red":     "FF3D5A",
    "blue":    "4D9EFF", "purple":  "B388FF", "text":    "E8F0FF",
    "dim":     "5A7090", "s1_bg":   "1A1200", "s2_bg":   "001A0A",
    "s3_bg":   "0A0014", "win_bg":  "0A1A0A", "loss_bg": "1A0A0A",
}

STRATEGY_META = {
    "S1_Asian_Breakout": ("FFB800", "S1 Asian Breakout"),
    "S2_Goldmine":       ("00E676", "S2 Goldmine"),
    "S3_Silver_Bullet":  ("B388FF", "S3 Silver Bullet"),
}

HEADERS = [
    "Date", "Time (IST)", "Strategy", "Direction", "Entry ($)",
    "SL ($)", "TP1 ($)", "TP2 ($)", "Size (oz)",
    "Exit ($)", "Gross P&L ($)", "Commission ($)", "Net P&L ($)",
    "Cumul. P&L ($)", "Equity ($)", "Result", "Notes"
]
WIDTHS = [12,10,20,10,10,10,10,10,9,10,13,12,12,14,14,8,22]


def _mk_wb():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XS
    from openpyxl.utils import get_column_letter

    wb   = Workbook()
    thin = XS(style="thin", color=C["border"])
    bdr  = lambda: Border(left=thin, right=thin, top=thin, bottom=thin)

    def fn(col, bold=False, sz=9):
        return Font(name="Consolas", bold=bold, size=sz, color=col)
    def fp(col): return PatternFill("solid", fgColor=col)
    def fa(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    def setup_sheet(ws, accent, tab_color):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab_color
        span = get_column_letter(len(HEADERS))

        ws.merge_cells(f"A1:{span}1")
        ws["A1"].value     = f"⚡  XAUUSD BOT — {ws.title.upper()}"
        ws["A1"].font      = fn(accent, True, 13)
        ws["A1"].fill      = fp(C["card"])
        ws["A1"].alignment = fa()
        ws.row_dimensions[1].height = 34

        ws.merge_cells(f"A2:{span}2")
        ws["A2"].value = "Asian Range Breakout | Goldmine | Silver Bullet | ATR SL | 1:2 R:R | Capital.com Demo"
        ws["A2"].font      = Font(name="Consolas", size=8, italic=True, color=C["dim"])
        ws["A2"].fill      = fp("04080F")
        ws["A2"].alignment = fa()
        ws.row_dimensions[2].height = 16

        for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
            c = ws.cell(row=3, column=i)
            c.value     = h
            c.font      = fn(accent, True, 9)
            c.fill      = fp(C["alt"])
            c.alignment = fa(wrap=True)
            c.border    = bdr()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[3].height = 28
        ws.freeze_panes = "A4"

    # All Trades
    ws_all = wb.active
    ws_all.title = "All Trades"
    setup_sheet(ws_all, C["gold"], C["gold"])

    # Strategy sheets
    for name, (accent, sheet_name) in STRATEGY_META.items():
        ws = wb.create_sheet(sheet_name)
        setup_sheet(ws, accent, accent)

    # Performance sheet
    ws_p = wb.create_sheet("Performance")
    ws_p.sheet_view.showGridLines = False
    ws_p.sheet_properties.tabColor = C["blue"]

    ws_p.merge_cells("A1:G1")
    ws_p["A1"].value     = "⚡  XAUUSD BOT — PERFORMANCE SUMMARY"
    ws_p["A1"].font      = fn(C["blue"], True, 13)
    ws_p["A1"].fill      = fp(C["card"])
    ws_p["A1"].alignment = fa()
    ws_p.row_dimensions[1].height = 34

    ph = [("Metric",28),("S1 Asian",16),("S2 Goldmine",16),("S3 Silver",16),("Combined",16),("Target",16),("Notes",28)]
    for i,(h,w) in enumerate(ph,1):
        c = ws_p.cell(row=2,column=i)
        c.value=h; c.font=fn(C["blue"],True,9); c.fill=fp(C["alt"])
        c.alignment=fa(wrap=True); c.border=bdr()
        ws_p.column_dimensions[get_column_letter(i)].width=w
    ws_p.row_dimensions[2].height=26

    rows = [
        ("Total Net P&L ($)",
         "=SUMIFS('S1 Asian Breakout'!M:M,'S1 Asian Breakout'!P:P,\"WIN\")+SUMIFS('S1 Asian Breakout'!M:M,'S1 Asian Breakout'!P:P,\"LOSS\")",
         "=SUMIFS('S2 Goldmine'!M:M,'S2 Goldmine'!P:P,\"WIN\")+SUMIFS('S2 Goldmine'!M:M,'S2 Goldmine'!P:P,\"LOSS\")",
         "=SUMIFS('S3 Silver Bullet'!M:M,'S3 Silver Bullet'!P:P,\"WIN\")+SUMIFS('S3 Silver Bullet'!M:M,'S3 Silver Bullet'!P:P,\"LOSS\")",
         "=B3+C3+D3","—","Net of commission"),
        ("Total Trades",
         "=COUNTA('S1 Asian Breakout'!A4:A10000)",
         "=COUNTA('S2 Goldmine'!A4:A10000)",
         "=COUNTA('S3 Silver Bullet'!A4:A10000)",
         "=B4+C4+D4","—",""),
        ("Winning Trades",
         "=COUNTIF('S1 Asian Breakout'!P:P,\"WIN\")",
         "=COUNTIF('S2 Goldmine'!P:P,\"WIN\")",
         "=COUNTIF('S3 Silver Bullet'!P:P,\"WIN\")",
         "=B5+C5+D5","—",""),
        ("Losing Trades",
         "=COUNTIF('S1 Asian Breakout'!P:P,\"LOSS\")",
         "=COUNTIF('S2 Goldmine'!P:P,\"LOSS\")",
         "=COUNTIF('S3 Silver Bullet'!P:P,\"LOSS\")",
         "=B6+C6+D6","—",""),
        ("Hit Ratio",
         "=IFERROR(B5/B4,0)","=IFERROR(C5/C4,0)","=IFERROR(D5/D4,0)",
         "=IFERROR(E5/E4,0)","S1:55% S2:82% S3:78%","Per research"),
        ("Avg Win ($)",
         "=IFERROR(AVERAGEIF('S1 Asian Breakout'!P:P,\"WIN\",'S1 Asian Breakout'!M:M),0)",
         "=IFERROR(AVERAGEIF('S2 Goldmine'!P:P,\"WIN\",'S2 Goldmine'!M:M),0)",
         "=IFERROR(AVERAGEIF('S3 Silver Bullet'!P:P,\"WIN\",'S3 Silver Bullet'!M:M),0)",
         "=IFERROR((B3+C3+D3)/(B5+C5+D5),0)","—",""),
        ("Avg Loss ($)",
         "=IFERROR(AVERAGEIF('S1 Asian Breakout'!P:P,\"LOSS\",'S1 Asian Breakout'!M:M),0)",
         "=IFERROR(AVERAGEIF('S2 Goldmine'!P:P,\"LOSS\",'S2 Goldmine'!M:M),0)",
         "=IFERROR(AVERAGEIF('S3 Silver Bullet'!P:P,\"LOSS\",'S3 Silver Bullet'!M:M),0)",
         "=IFERROR((B3+C3+D3)/(B6+C6+D6),0)","—",""),
        ("Gain:Loss Ratio",
         "=IFERROR(ABS(B7/B8),0)","=IFERROR(ABS(C7/C8),0)","=IFERROR(ABS(D7/D8),0)",
         "=IFERROR(ABS(E7/E8),0)",">1.5x","Target minimum"),
        ("Max Win ($)",
         "=IFERROR(MAXIFS('S1 Asian Breakout'!M:M,'S1 Asian Breakout'!P:P,\"WIN\"),0)",
         "=IFERROR(MAXIFS('S2 Goldmine'!M:M,'S2 Goldmine'!P:P,\"WIN\"),0)",
         "=IFERROR(MAXIFS('S3 Silver Bullet'!M:M,'S3 Silver Bullet'!P:P,\"WIN\"),0)",
         "=MAX(B10,C10,D10)","—",""),
        ("Max Loss ($)",
         "=IFERROR(MINIFS('S1 Asian Breakout'!M:M,'S1 Asian Breakout'!P:P,\"LOSS\"),0)",
         "=IFERROR(MINIFS('S2 Goldmine'!M:M,'S2 Goldmine'!P:P,\"LOSS\"),0)",
         "=IFERROR(MINIFS('S3 Silver Bullet'!M:M,'S3 Silver Bullet'!P:P,\"LOSS\"),0)",
         "=MIN(B11,C11,D11)","—",""),
    ]

    for r, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            c = ws_p.cell(row=r, column=ci, value=val)
            c.fill      = fp(C["alt"] if r%2==0 else C["card"])
            c.alignment = fa() if ci>1 else Alignment(horizontal="left", vertical="center")
            c.border    = bdr()
            c.font      = fn(C["text"], ci==1, 9)
        ws_p.row_dimensions[r].height = 22

    for col in (2,3,4,5):
        ws_p.cell(row=7, column=col).number_format = "0.0%"

    wb.save(EXCEL_PATH)
    log.info(f"✅ XAUUSD Excel workbook created: {EXCEL_PATH}")


def _ensure_wb():
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_PATH):
        _mk_wb()
        return load_workbook(EXCEL_PATH)
    try:
        wb = load_workbook(EXCEL_PATH)
        required = ["All Trades","S1 Asian Breakout","S2 Goldmine","S3 Silver Bullet","Performance"]
        for s in required:
            if s not in wb.sheetnames:
                raise ValueError(f"Missing: {s}")
        return wb
    except Exception as e:
        log.warning(f"Excel invalid ({e}) — recreating")
        try: os.remove(EXCEL_PATH)
        except: pass
        _mk_wb()
        return load_workbook(EXCEL_PATH)


def log_trade(trade: dict, account_equity: float = 0, cumulative_pnl: float = 0):
    """Thread-safe — never raises."""
    with _lock:
        try:
            _write_row(trade, account_equity, cumulative_pnl)
        except Exception as e:
            log.error(f"Excel write error: {e}")


def _write_row(trade: dict, equity: float, cum_pnl: float):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XS

    wb = _ensure_wb()

    strategy  = str(trade.get("strategy", ""))
    direction = str(trade.get("direction", ""))
    entry     = float(trade.get("entry",  0) or 0)
    sl        = float(trade.get("sl",     0) or 0)
    tp1       = float(trade.get("tp1",    0) or 0)
    tp2       = float(trade.get("tp2",    0) or 0)
    size      = float(trade.get("size",   0) or 0)
    exit_p    = float(trade.get("exit_price", 0) or 0)
    gross     = trade.get("gross_pnl")
    result    = str(trade.get("result",  ""))
    reason    = str(trade.get("reason",  ""))
    ts_raw    = trade.get("time", "")

    try:
        ts = datetime.fromisoformat(str(ts_raw).replace("Z","+00:00"))
        if ts.tzinfo is None:
            ts = IST.localize(ts)
        ts_ist = ts.astimezone(IST)
    except Exception:
        ts_ist = datetime.now(IST)

    date_str = ts_ist.strftime("%Y-%m-%d")
    time_str = ts_ist.strftime("%H:%M:%S")

    gross_v  = float(gross) if gross is not None else None
    comm     = round(size * COMMISSION, 2)
    net_pnl  = round(gross_v - comm, 2) if gross_v is not None else None

    # Row color
    if result == "WIN":   bg = C["win_bg"]
    elif result == "LOSS":bg = C["loss_bg"]
    elif "S1" in strategy:bg = C["s1_bg"]
    elif "S2" in strategy:bg = C["s2_bg"]
    elif "S3" in strategy:bg = C["s3_bg"]
    else:                  bg = C["card"]

    thin   = XS(style="thin", color=C["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_vals = [
        date_str, time_str,
        strategy.replace("_"," "),
        direction,
        entry if entry else "",
        sl    if sl    else "",
        tp1   if tp1   else "",
        tp2   if tp2   else "",
        size  if size  else "",
        exit_p if exit_p else "",
        gross_v if gross_v is not None else "",
        comm,
        net_pnl if net_pnl is not None else "",
        round(cum_pnl, 2),
        round(equity,  2),
        result,
        reason,
    ]

    STRAT_SHEET = {
        "S1_Asian_Breakout": "S1 Asian Breakout",
        "S2_Goldmine":       "S2 Goldmine",
        "S3_Silver_Bullet":  "S3 Silver Bullet",
    }
    target_sheets = ["All Trades"]
    if strategy in STRAT_SHEET:
        target_sheets.append(STRAT_SHEET[strategy])

    accent = STRATEGY_META.get(strategy, (C["gold"], ""))[0]

    for sname in target_sheets:
        if sname not in wb.sheetnames:
            continue
        ws       = wb[sname]
        next_row = max(ws.max_row + 1, 4)

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=next_row, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border

            if ci in (11, 13, 14) and val != "":
                try:
                    col = C["green"] if float(val) >= 0 else C["red"]
                    cell.font = Font(name="Consolas", size=9, bold=True, color=col)
                except Exception:
                    cell.font = Font(name="Consolas", size=9, color=C["text"])
            elif ci == 3:  # Strategy
                cell.font = Font(name="Consolas", size=9, bold=True, color=accent)
            elif ci == 4:  # Direction
                col = C["green"] if direction=="BUY" else C["red"]
                cell.font = Font(name="Consolas", size=9, bold=True, color=col)
            elif ci == 16:  # Result
                col = C["green"] if result=="WIN" else C["red"] if result=="LOSS" else C["dim"]
                cell.font = Font(name="Consolas", size=9, bold=True, color=col)
            elif ci == 17:  # Notes
                cell.font = Font(name="Consolas", size=8, italic=True, color=C["dim"])
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = Font(name="Consolas", size=9, color=C["text"])

        ws.row_dimensions[next_row].height = 20

    wb.save(EXCEL_PATH)
    log.info(f"Excel ✅ {strategy} {direction} | net=${net_pnl} | result={result}")
